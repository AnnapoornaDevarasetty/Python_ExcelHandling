"""
 * @file excel_project.py
 * @author Devarasetty V S Annapoorna
 * @brief To read and store the data in excel files
 * @version 3.8.5
 * @date 2021-07-06

"""
import openpyxl


class SemesterMarks:
    """
    Class for the getting the Semester Marks of the given PS Number
    """
    def __init__(self, sheet_name, ps_number, sheets, excel_document):
        """
        To initialize the attributes of the class SemesterMarks
        :param sheet_name:Name of the Sheet given by the user
        :param ps_number: The PS Number of the person to access the data
        :param sheets: To access the sheets in the excel
        :param excel_document: To get the excel file path
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number
        self.sheets = sheets
        self.excel_document = excel_document

    def semester_fun(self):
        """
        Method semester_fun is to get the semester marks data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            # Path to store the new excel file
            filepath = "C:\\Users\\Mallikarjuna rao\\Desktop\\NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion=input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = self.excel_document.get_sheet_by_name("Semester_Marks")
            sheet_obj = self.excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")


class HobbiesList:
    """
    Class for the getting the Hobbies List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number, sheets, excel_document):
        """
        To initialize the attributes of the class HobbiesList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        :param sheets:To access the sheets in the excel
        :param excel_document: To get the excel file path
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number
        self.sheets = sheets
        self.excel_document = excel_document

    def hobbies_fun(self):
        """
        Method hobbies_fun is to get the hobbies data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            # Path to store the new excel file
            filepath = "C:\\Users\\Mallikarjuna rao\\Desktop\\NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = self.excel_document.get_sheet_by_name("Hobbies_List")
            sheet_obj = self.excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")


class CitiesVisited:
    """
    Class for the getting the Cities Visited List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number, sheets, excel_document):
        """
        To initialize the attributes of the class CitiesVisitedList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        :param sheets:To access the sheets in the excel
        :param excel_document: To get the excel file path
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number
        self.sheets = sheets
        self.excel_document = excel_document

    def cities_fun(self):
        """
        Method cities_fun is to get the cities visited data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            # Path to store the new excel file
            filepath = "C:\\Users\\Mallikarjuna rao\\Desktop\\NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = self.excel_document.get_sheet_by_name("Cities_Visited")
            sheet_obj = self.excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")


class ProgrammingLanguage:
    """
    Class for the getting the Programming Language Expertise List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number, sheets, excel_document):
        """
        To initialize the attributes of the class ProgrammingLanguageList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        :param sheets:To access the sheets in the excel
        :param excel_document: To get the excel file path
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number
        self.sheets = sheets
        self.excel_document = excel_document

    def programming_fun(self):
        """
        Method programming_fun is to get the programming expertise data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            # Path to store the new excel file
            filepath = "C:\\Users\\Mallikarjuna rao\\Desktop\\NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = self.excel_document.get_sheet_by_name("ProgrammingLanguage_Expertise")
            sheet_obj = self.excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")


class SportsList:
    """
    Class for the getting the Sports List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number, sheets, excel_document):
        """
        To initialize the attributes of the class HobbiesList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        :param sheets:To access the sheets in the excel
        :param excel_document: To get the excel file path
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number
        self.sheets = sheets
        self.excel_document = excel_document

    def sports_fun(self):
        """
        Method sports_fun is to get the sports data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            # Path to store the new excel file
            filepath = "C:\\Users\\Mallikarjuna rao\\Desktop\\NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = self.excel_document.get_sheet_by_name("Sports_List")
            sheet_obj = self.excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")


class Main:
    """
    Main class to call all the methods
    """
    def __init__(self):
        """
        To initialize the attributes of the Main class
        """
        # Opening the Excel file in which the data is available
        excel_document = openpyxl.load_workbook("Data.xlsx")
        sheet = excel_document.get_sheet_by_name("Semester_Marks")
        sheet_obj = excel_document.active
        max_row = sheet_obj.max_row
        flag = 0
        ps_number_list = []
        print("\nThe PS Numbers present in excel sheet are : ")
        for i in range(2, max_row + 1):
            cell_obj = sheet.cell(row=i, column=1)
            ps_number_list.append(cell_obj.value)
            print(str(i - 1) + "." + str(cell_obj.value))
        print("\nEnter a PS number from the above list: ")
        ps_number = int(input())
        for number in range(0, len(ps_number_list)):
            if ps_number == ps_number_list[number]:
                flag = 1
        if flag == 0:
            print("Please try to enter valid PS Number")
            exit()
        # Getting the sheet names
        print("\nThe Categories present in the excel file are : \n")
        sheets = excel_document.sheetnames
        for category in range(0, len(sheets)):
            print(str(category + 1) + "." + sheets[category])
        # Entering the Category Name to get a particular Data
        print("\nEnter Category name for which you want details : ")
        sheet_name = input()
        if sheet_name == "Semester_Marks":
            semester_marks_ob = SemesterMarks(sheet_name, ps_number, sheets, excel_document)
            semester_marks_ob.semester_fun()
        elif sheet_name == "Hobbies_List":
            hobbies_list_ob = HobbiesList(sheet_name, ps_number, sheets, excel_document)
            hobbies_list_ob.hobbies_fun()
        elif sheet_name == "Cities_Visited":
            cities_visited_ob = CitiesVisited(sheet_name, ps_number, sheets, excel_document)
            cities_visited_ob.cities_fun()
        elif sheet_name == "ProgrammingLanguage_Expertise":
            programming_language_ob = ProgrammingLanguage(sheet_name, ps_number, sheets, excel_document)
            programming_language_ob.programming_fun()
        elif sheet_name == "Sports_List":
            sports_list_ob = SportsList(sheet_name, ps_number, sheets, excel_document)
            sports_list_ob.sports_fun()
        else:
            print("The sheet name you entered is not available..Please try again...")


main_ob = Main()
