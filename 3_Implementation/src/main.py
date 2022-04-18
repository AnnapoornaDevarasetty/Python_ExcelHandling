import openpyxl

from excel_project import *

class Main:
    """
    Main class to call all the methods
    """
    def mainFuntion(self):
        """
        To initialize the attributes of the Main class
        """
        # Opening the Excel file in which the data is available
        excel_document = openpyxl.load_workbook("Data.xlsx")
        sheet = excel_document.get_sheet_by_name("Semester_Marks")
        sheet_obj = excel_document.active
        max_row = sheet_obj.max_row
        flag = 0
        ps_number_list = []
        print("\nThe PS Numbers present in excel sheet are : ")
        for i in range(2, max_row + 1):
            cell_obj = sheet.cell(row=i, column=1)
            ps_number_list.append(cell_obj.value)
            print(str(i - 1) + "." + str(cell_obj.value))
        print("\nEnter a PS number from the above list: ")
        ps_number = int(input())
        for number in range(0, len(ps_number_list)):
            if ps_number == ps_number_list[number]:
                flag = 1
        if flag == 0:
            print("Please try to enter valid PS Number")
            exit()
        # Getting the sheet names
        print("\nThe Categories present in the excel file are : \n")
        sheets = excel_document.sheetnames
        for category in range(0, len(sheets)):
            print(str(category + 1) + "." + sheets[category])
        # Entering the Category Name to get a particular Data
        print("\nEnter Category name for which you want details : ")
        sheet_name = input()
        filepath = "NewData.xlsx"
        work_book = openpyxl.load_workbook(filepath)
        updated_sheet = work_book.active
        if sheet_name == "Semester_Marks":
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
                work_book.save(filepath)
            semester_marks_ob = SemesterMarks(sheet_name, ps_number)
            val=semester_marks_ob.semester_fun()
        elif sheet_name == "Hobbies_List":
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
                work_book.save(filepath)
            hobbies_list_ob = HobbiesList(sheet_name, ps_number)
            hobbies_list_ob.hobbies_fun()
        elif sheet_name == "Cities_Visited":
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
                work_book.save(filepath)
            cities_visited_ob = CitiesVisited(sheet_name, ps_number)
            cities_visited_ob.cities_fun()
        elif sheet_name == "ProgrammingLanguage_Expertise":
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
                work_book.save(filepath)
            programming_language_ob = ProgrammingLanguage(sheet_name, ps_number)
            programming_language_ob.programming_fun()
        elif sheet_name == "Sports_List":
            print("\nDo you want to delete the previous data(Yes/No): ")
            opinion = input()
            if opinion.lower() == 'yes' or opinion.lower() == 'y':
                updated_sheet.delete_rows(1, updated_sheet.max_row)
                work_book.save(filepath)
            sports_list_ob = SportsList(sheet_name, ps_number)
            sports_list_ob.sports_fun()
        else:
            print("The sheet name you entered is not available..Please try again...")


main_ob = Main()
main_ob.mainFuntion()