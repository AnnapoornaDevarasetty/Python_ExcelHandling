"""
 * @file excel_project.py
 * @author Devarasetty V S Annapoorna
 * @brief To read and store the data in excel files
 * @version 3.8.5
 * @date 2021-07-06

"""
import openpyxl


class SemesterMarks:
    """
    Class for the getting the Semester Marks of the given PS Number
    """
    def __init__(self, sheet_name, ps_number):
        """
        To initialize the attributes of the class SemesterMarks
        :param sheet_name:Name of the Sheet given by the user
        :param ps_number: The PS Number of the person to access the data
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number

    def semester_fun(self):
        """
        Method semester_fun is to get the semester marks data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            excel_document = openpyxl.load_workbook("3_Implementation/src/Data.xlsx")
            # Path to store the new excel file
            filepath = "3_Implementation/src/NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active

            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = excel_document.get_sheet_by_name("Semester_Marks")
            sheet_obj = excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")
        finally:
            return flag


class HobbiesList:
    """
    Class for the getting the Hobbies List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number):
        """
        To initialize the attributes of the class HobbiesList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number

    def hobbies_fun(self):
        """
        Method hobbies_fun is to get the hobbies data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            excel_document = openpyxl.load_workbook("3_Implementation/src/Data.xlsx")
            # Path to store the new excel file
            filepath = "3_Implementation/src/NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = excel_document.get_sheet_by_name("Hobbies_List")
            sheet_obj = excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")
        finally:
            return flag


class CitiesVisited:
    """
    Class for the getting the Cities Visited List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number):
        """
        To initialize the attributes of the class CitiesVisitedList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number

    def cities_fun(self):
        """
        Method cities_fun is to get the cities visited data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            excel_document = openpyxl.load_workbook("3_Implementation/src/Data.xlsx")
            # Path to store the new excel file
            filepath = "3_Implementation/src/NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = excel_document.get_sheet_by_name("Cities_Visited")
            sheet_obj = excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")
        finally:
            return flag


class ProgrammingLanguage:
    """
    Class for the getting the Programming Language Expertise List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number):
        """
        To initialize the attributes of the class ProgrammingLanguageList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number

    def programming_fun(self):
        """
        Method programming_fun is to get the programming expertise data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            excel_document = openpyxl.load_workbook("3_Implementation/src/Data.xlsx")
            # Path to store the new excel file
            filepath = "3_Implementation/src/NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = excel_document.get_sheet_by_name("ProgrammingLanguage_Expertise")
            sheet_obj = excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")
        finally:
            return flag


class SportsList:
    """
    Class for the getting the Sports List of the given PS Number
    """
    def __init__(self, sheet_name, ps_number):
        """
        To initialize the attributes of the class HobbiesList
        :param sheet_name: Name of the Sheet given by the user
        :param ps_number:The PS Number of the person to access the data
        """
        self.sheet_name = sheet_name
        self.ps_number = ps_number

    def sports_fun(self):
        """
        Method sports_fun is to get the sports data of requested ps number
        :return:The Requested data will be entered into a file NewData.xlsx
        """
        try:
            excel_document = openpyxl.load_workbook("3_Implementation/src/Data.xlsx")
            # Path to store the new excel file
            filepath = "3_Implementation/src/NewData.xlsx"
            work_book = openpyxl.load_workbook(filepath)
            updated_sheet = work_book.active
            # Creating a tuple filtered values to enter the data
            filtered_values = ()
            # Titles tuple is to store the sheet names
            titles = ()
            flag = 0
            sheet = excel_document.get_sheet_by_name("Sports_List")
            sheet_obj = excel_document.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            for j in range(1, max_col + 1):
                cell_column = sheet.cell(row=1, column=j)
                titles = titles + (cell_column.value,)
            updated_sheet.append(titles)
            # Will print row value according to the PS Number
            for i in range(2, max_row + 1):
                cell_obj = sheet.cell(row=i, column=1)
                if self.ps_number == cell_obj.value:
                    flag = 1
                    for j in range(1, max_col + 1):
                        cell_obj = sheet.cell(row=i, column=j)
                        filtered_values = filtered_values + (cell_obj.value,)
            if flag == 0:
                print("\nThe PS Number you entered is not valid.... Please try again...")
            else:
                updated_sheet.append(filtered_values)
                work_book.save(filepath)
                print("\nThe data you have requested is added in this NewData.xlsx sheet\n")
        except:
            print("Unable to complete the process...Please try again...\n")
        finally:
            return flag




